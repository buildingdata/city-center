"""Extract city center coordinates from a PDF and export them to Excel."""

from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent
PDF_PATH = ROOT / "cityCenter.pdf"
OUTPUT_PATH = ROOT / "city_center_output.xlsx"
HEADERS = ["城市", "经度", "纬度", "瓦片", "UTM"]
FONT_NAME = "等线"
TARGET_CITIES = [
    "哈尔滨",
    "漠河",
    "富蕴",
    "佳木斯",
    "长春",
    "呼和浩特",
    "沈阳",
    "西宁",
    "西安",
    "拉萨",
    "喀什",
    "北京",
    "兰州",
    "青岛",
    "银川",
    "石家庄",
    "吐鲁番",
    "重庆",
    "武汉",
    "合肥",
    "长沙",
    "南昌",
    "上海",
    "南京",
    "杭州",
    "广州",
    "海口",
    "三亚",
    "福州",
    "香港",
    "澳门",
    "台北",
    "昆明",
]
TARGET_ALIASES = {
    "喀什": "喀什地区",
    "吐鲁番": "吐鲁番地区",
}


@dataclass(frozen=True)
class CityRecord:
    """A normalized city row with derived tile and UTM values."""

    city: str
    longitude: float
    latitude: float
    tile: str
    utm: str

    def as_row(self) -> list[object]:
        """Return the record as an Excel row."""

        return [self.city, self.longitude, self.latitude, self.tile, self.utm]


def normalize_pdf_text(pdf_path: Path) -> str:
    """Read the PDF text layer and remove all whitespace for easier parsing."""

    reader = PdfReader(str(pdf_path))
    text = "".join(page.extract_text() or "" for page in reader.pages)
    return re.sub(r"\s+", "", text)


def parse_coordinate_pair(geo_value: str) -> tuple[float, float]:
    """Parse a longitude/latitude pair from the PDF geo field."""

    lon_text, lat_text = geo_value.split("|", 1)[0].split(",", 1)
    return float(lon_text), float(lat_text)


def format_utm(longitude: float, latitude: float) -> str:
    """Convert longitude and latitude to a UTM zone label like `49N`."""

    zone = int(math.floor((longitude + 180.0) / 6.0) + 1)
    zone = max(1, min(60, zone))
    hemisphere = "N" if latitude >= 0 else "S"
    return f"{zone}{hemisphere}"


def format_lon_boundary(value: int) -> str:
    """Format a longitude tile boundary with a 3-digit absolute value."""

    direction = "e" if value >= 0 else "w"
    return f"{direction}{abs(value):03d}"


def format_lat_boundary(value: int) -> str:
    """Format a latitude tile boundary with a 2-digit absolute value."""

    direction = "n" if value >= 0 else "s"
    return f"{direction}{abs(value):02d}"


def format_tile(longitude: float, latitude: float) -> str:
    """Build the required 5° x 5° tile name for a coordinate pair."""

    lon_min = math.floor(longitude / 5.0) * 5
    lon_max = lon_min + 5
    lat_min = math.floor(latitude / 5.0) * 5
    lat_max = lat_min + 5
    return "_".join(
        [
            format_lon_boundary(int(lon_min)),
            format_lat_boundary(int(lat_max)),
            format_lon_boundary(int(lon_max)),
            format_lat_boundary(int(lat_min)),
        ]
    )


def build_record(city_name: str, geo_value: str) -> CityRecord:
    """Create a normalized city record from the raw PDF values."""

    longitude, latitude = parse_coordinate_pair(geo_value)
    return CityRecord(
        city=city_name,
        longitude=longitude,
        latitude=latitude,
        tile=format_tile(longitude, latitude),
        utm=format_utm(longitude, latitude),
    )


def extract_pdf_records(pdf_path: Path) -> list[CityRecord]:
    """Extract all city records from the PDF text structure."""

    text = normalize_pdf_text(pdf_path)
    municipalities_match = re.search(r"municipalities:\[(.*?)\],provinces:", text)
    provinces_match = re.search(r"provinces:\[(.*)\],other:\[", text)
    other_match = re.search(r"other:\[(.*?)\]\};?$", text)

    if municipalities_match is None or provinces_match is None or other_match is None:
        raise ValueError("无法从 PDF 文本中识别城市数据结构。")

    object_pattern = re.compile(r'\{n:"([^"]+)",g:"([^"]+)"\}')
    province_pattern = re.compile(r'\{n:"([^"]+)",g:"([^"]+)",cities:\[(.*?)\]\}')

    records: list[CityRecord] = []

    for city_name, geo_value in object_pattern.findall(municipalities_match.group(1)):
        records.append(build_record(city_name, geo_value))

    for _, _, cities_blob in province_pattern.findall(provinces_match.group(1)):
        for city_name, geo_value in object_pattern.findall(cities_blob):
            records.append(build_record(city_name, geo_value))

    for city_name, geo_value in object_pattern.findall(other_match.group(1)):
        records.append(build_record(city_name, geo_value))

    return records


def autosize_columns(worksheet) -> None:
    """Adjust worksheet column widths based on their longest displayed value."""

    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_width = max(len(value) for value in values)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_width + 2


def style_sheet(worksheet) -> None:
    """Apply the requested font and number formatting to a worksheet."""

    font = Font(name=FONT_NAME)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = font
            if cell.row > 1 and cell.column in (2, 3) and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"
    autosize_columns(worksheet)


def write_sheet(worksheet, rows: list[list[object]]) -> None:
    """Write headers and data rows to a worksheet, then style it."""

    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    style_sheet(worksheet)


def build_sheet1_rows(records: list[CityRecord]) -> list[list[object]]:
    """Build the requested Sheet1 rows in the user-provided city order."""

    record_lookup = {record.city: record for record in records}
    rows: list[list[object]] = []

    for target_city in TARGET_CITIES:
        lookup_name = TARGET_ALIASES.get(target_city, target_city)
        record = record_lookup.get(lookup_name)
        if record is None:
            rows.append([target_city, "", "", "", ""])
            continue
        rows.append([target_city, record.longitude, record.latitude, record.tile, record.utm])

    return rows


def write_workbook(
    sheet1_rows: list[list[object]],
    sheet2_rows: list[list[object]],
    output_path: Path,
) -> None:
    """Create the workbook with Sheet1 and Sheet2 and save it to disk."""

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet2 = workbook.create_sheet("Sheet2")

    write_sheet(sheet1, sheet1_rows)
    write_sheet(sheet2, sheet2_rows)
    workbook.save(output_path)


def main() -> int:
    """Run the PDF extraction pipeline and write the output workbook."""

    if not PDF_PATH.exists():
        raise FileNotFoundError(f"未找到 PDF 文件: {PDF_PATH}")

    records = extract_pdf_records(PDF_PATH)
    sheet1_rows = build_sheet1_rows(records)
    sheet2_rows = [record.as_row() for record in records]
    write_workbook(sheet1_rows, sheet2_rows, OUTPUT_PATH)

    missing_cities = [row[0] for row in sheet1_rows if row[1] == ""]
    print(f"已生成: {OUTPUT_PATH.name}")
    print(f"Sheet2 城市数: {len(sheet2_rows)}")
    if missing_cities:
        print("Sheet1 留空城市: " + "、".join(missing_cities))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
